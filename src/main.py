print("=== IMPORTING LIBRARIES ===")
import time
import pandas as pd
import os
import platform
import datetime
from datetime import date
import json
import numpy as np
from datetime import date
from appJar import gui
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
from openpyxl.chart.series import DataPoint
from openpyxl.chart import (
    PieChart,
    Reference
)
from openpyxl.styles import (
    Border,
    Font,
    Side
)

# import numpy.random.common
# import numpy.random.bounded_integers
# import numpy.random.entropy

print("=== FINISHED IMPORTING LIBRARIES ===")


# Creating the Gui Interface
class Gui:

    # Creating the variables this class with execute at the beginning each time
    def __init__(self):
        self.email = ""
        self.password = ""
        self.output = ""
        self.gui = gui("Ken's Report Program", "780x240")
        self.gui.setResizable(canResize=False)
        self.gui.setLocation("CENTER")
        self.gui.setIcon("../Pictures/hh.ico")
        self.gui.setBg("#2A9AA9")  # Housing Hand Blue
        self.get_excel_file()
        self.new_pass()
        self.button()
        self.start_gui()

    def new_pass(self):
        self.gui.addLabel("Password", "Have you changed your password?", 1, 2, 3, 0)
        self.gui.addRadioButton("password", "Yes", 2, 3, 0, 0)
        self.gui.addRadioButton("password", "No", 2, 4, 0, 0)
        self.gui.setRadioButton("password", "Yes")
        self.gui.setRadioButtonFg("password", "#333333")
        self.gui.setLabelFg("Password", "#333333")

    # Running the Gui
    def start_gui(self):
        self.gui.go()

    # The Input of the file
    def get_excel_file(self):
        self.gui.addFileEntry("excel_file", 0, 2, 3, 1)
        self.gui.setEntryBg("excel_file", "#ffb701")  # Housing Hand Orange
        self.gui.setEntryDefault("excel_file", "Please Enter The Excel File")

    # The function that runs when the Button is pressed
    def excel_output(self, button):
        if button == "Enter":
            self.output = self.gui.getEntry("excel_file")
            if self.gui.getRadioButton('password') == 'Yes':
                self.gui.stop()
                CredentialsGui()
            else:
                pass
            if self.output == "":
                self.gui.errorBox("Error", "Please select a File.")
            else:
                self.gui.stop()
            return self.output

    def button(self):  # The Enter button
        self.gui.addButton("Enter", self.excel_output, 4, 2, 3, 0)


class CredentialsGui:

    def __init__(self):
        self.gui = gui("Ken's Report Program", "780x240")
        self.gui.setResizable(canResize=False)
        self.gui.setLocation("CENTER")
        self.gui.setIcon("../Pictures/hh.ico")
        self.password_file = "../Credentials"
        self.cred_json = ""
        self.email = ""
        self.password = ""
        self.data = {}
        self.gui.setBg("#2A9AA9")
        self.get_credentials()
        self.buttons()
        self.check_data()
        self.start_gui()

    def write_data(self, email, password):
        self.data["Credentials"].clear()
        self.data["Credentials"].append(
            {
                "Email": email,
                "Password": password
            }
        )

        with open(self.password_file + "/credentials.ko", "w") as cred_json:
            json.dump(self.data, cred_json, sort_keys=True, indent=4, separators=(',', ': '))

    def check_data(self):
        if os.path.exists(self.password_file) is False:
            os.mkdir("Credentials")
            self.data = {"Credentials": []}
            with open(self.password_file + '/credentials.ko', 'w') as self.cred_json:
                json.dump(self.data, self.cred_json, sort_keys=True, indent=4, separators=(',', ': '))

        elif (os.path.exists(self.password_file)) & (os.path.exists(self.password_file + "/credentials.ko") is False):
            self.data = {"Credentials": []}
            with open(self.password_file + '/credentials.ko', 'w') as self.cred_json:
                json.dump(self.data, self.cred_json, sort_keys=True, indent=4, separators=(',', ': '))

        elif (os.path.exists(self.password_file)) & (os.path.exists(self.password_file + "/credentials.ko")):
            self.data = {"Credentials": []}
            with open(self.password_file + '/credentials.ko', 'w') as self.cred_json:
                json.dump(self.data, self.cred_json, indent=4, separators=(',', ': '))
            print("Done")

    def get_credentials(self):
        self.gui.addLabel("email", "Email", 0, 0)
        self.gui.addEntry("email_entry", 0, 1)
        self.gui.addLabel("password", "Password", 1, 0)
        self.gui.addSecretEntry("password_entry", 1, 1)
        self.gui.setLabelFg("email", "ffb701")
        self.gui.setLabelFg("password", "ffb701")

    def press(self, button):
        if button == "Cancel":
            self.gui.stop()

        elif button == "Clear":
            if len(self.gui.getEntry("email_entry")) == 0 or len(self.gui.getEntry("password_entry")) == 0:
                self.gui.errorBox("Error", "You must enter your Email or Password")
                self.gui.setFocus("email_entry")
            else:
                self.gui.clearEntry("email_entry")
                self.gui.clearEntry("password_entry")
                self.gui.setFocus("email_entry")

        elif button == "Enter":
            if (len(self.gui.getEntry("email_entry")) == 0) | (len(self.gui.getEntry("password_entry")) == 0):
                self.gui.errorBox("Error", "You must enter your Email or Password")
                self.gui.setFocus("email_entry")
            if (len(self.gui.getEntry("email_entry")) > 0) & (len(self.gui.getEntry("password_entry")) >= 8):
                self.email = self.gui.getEntry("email_entry")
                self.password = self.gui.getEntry("password_entry")
                self.write_data(self.email, self.password)
                self.gui.stop()
            elif (len(self.gui.getEntry("password_entry")) < 8) & (len(self.gui.getEntry("email_entry")) > 0):
                self.gui.errorBox("Error", "Invalid credentials \n"
                                           "(Please check if your password and email are correct)")

    def buttons(self):
        self.gui.addButtons(["Clear", "Enter", "Cancel"], self.press, 2, 0, 4)

    def start_gui(self):
        self.gui.go()


# This is the robot that is crawling the website
class Applicants:

    # Creating the variables this class with execute at the beginning each time
    def __init__(self, username, pw):
        self.client_journey = []
        self.href_links = []
        self.client_names = []
        self.nationality, self.accomodation_provider, self.start_date, self.status, self.sales_source, self.affiliates, self.fast_track = [], [], [], [], [], [], []
        self.date_created = []
        self.duplicates = []
        self.affiliates_code = []
        self.seconds = 2.10
        self.search_bar = ""
        self.links = ""
        self.quote_button = ""
        self.applicant_link = "/dashboard/profile/"
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get("https:panacea.housinghand.co.uk")
        self.usernamebox = self.driver.find_element_by_id("emailAddress")
        self.usernamebox.send_keys(username)
        self.passwordbox = self.driver.find_element_by_id("password")
        self.passwordbox.send_keys(pw + Keys.RETURN)
        time.sleep(self.seconds)
        self.applicants_button = self.driver.find_element_by_xpath('//a[@href = "/dashboard/applicant"]')
        self.os_base = platform.system()

    # Pressing the Warning Button when available.
    def warning_button(self):
        try:
            self.driver.find_element_by_xpath('//button[@class = "btn btn-warning"]').click()
        except Exception:
            pass

    # extracting the data from each applicant
    def get_data(self, ref_number):
        self.applicants_button.click()
        time.sleep(self.seconds)
        self.search_bar = self.driver.find_element_by_class_name("search-box")

        for i in range(0, len(ref_number)):
            self.search_bar.send_keys(ref_number[i][1])
            time.sleep(self.seconds)
            self.store_links()
            time.sleep(self.seconds)
            self.get_info(i)
            time.sleep(self.seconds)
            self.search_bar = self.driver.find_element_by_class_name("search-box")
            if self.os_base != "Darwin":
                self.search_bar.send_keys(Keys.CONTROL, "a")
                self.search_bar.send_keys(Keys.DELETE)
            else:
                self.search_bar.clear()

            # Looking for duplicates and cleaning the list in order to get away with crashes cause from duplicates
            if len(self.href_links) == 1:
                self.href_links.pop()
            elif len(self.href_links) > 1:
                self.href_links.clear()
                self.duplicates.append(ref_number[i][1])
            time.sleep(self.seconds)
        self.driver.quit()

        return self.nationality, self.accomodation_provider, self.start_date, self.duplicates, self.status, \
               self.fast_track

    # Storing the url id of each applicant
    def store_links(self):
        self.links = self.driver.find_elements_by_xpath(" //a [@class= 'card__link' ] ")

        for link in self.links:
            link = link.get_attribute("href")
            link = link.split("/")
            self.href_links.append(link[5])
        return self.href_links

    # Using the url id retrieved above we get inside each applicant
    def get_info(self, x):
        try:
            self.driver.find_element_by_xpath("//a[@href ='" + self.applicant_link + self.href_links[0] + "']" +
                                              '//*[@class = "card card--applicant col-xs-12 col-sm-12 col-md-6 col-lg-4"]').click()
        except NoSuchElementException as e:
            print(e)

        time.sleep(self.seconds)
        self.warning_button()

        if len(self.nationality) == 0:
            self.nationality.append(self.driver.find_element_by_xpath
                                    ('//app-inline-edit[@label = "Nationality"]').text)
        else:
            self.nationality.append(self.driver.find_element_by_xpath
                                    ('//app-inline-edit[@label = "Nationality"]').text)

        if len(self.status) == 0:
            self.status.append(self.driver.find_elements_by_xpath
                               ('//div[@class = "card__status-text ml-1"]')[1].text)
        else:
            self.status.append(self.driver.find_elements_by_xpath
                               ('//div[@class = "card__status-text ml-1"]')[1].text)

        self.affiliates_code = self.driver.find_elements_by_xpath(
            '//div[@class = "col-xs-5 col-sm-3 col-md-5 col-lg-5 font-opacity-2"] ')[3].text
        # print(self.affiliates_code)
        if self.affiliates_code == "Affiliate Code":
            self.affiliates.append(self.driver.find_elements_by_xpath(
                '//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"] ')[4].text)

            if self.driver.find_element_by_xpath(
                    '//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7 font-blue-02"]').text == "Student":
                if len(self.sales_source) == 0:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[8].text)
                else:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[8].text)
            elif (self.driver.find_element_by_xpath(
                    '//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7 font-blue-02"]').text) == "Working Professional":

                if len(self.sales_source) == 0:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[7].text)
                else:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[7].text)
        else:
            if self.driver.find_element_by_xpath(
                    '//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7 font-blue-02"]').text == "Student":
                if len(self.sales_source) == 0:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[7].text)
                else:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[7].text)
            elif self.driver.find_element_by_xpath(
                    '//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7 font-blue-02"]').text == "Working Professional":
                if len(self.sales_source) == 0:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[6].text)
                else:
                    self.sales_source.append(self.driver.find_elements_by_xpath
                                             ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[6].text)

        # /*********************************Here I will insert the code for the fast_track applicants.*******************************************************/
        if len(self.fast_track) == 0:
            self.fast_track.append(self.driver.find_elements_by_xpath
                                   ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[-1].text)
        else:
            self.fast_track.append(self.driver.find_elements_by_xpath
                                   ('//div[@class = "col-xs-7 col-sm-9 col-md-7 col-lg-7"]')[-1].text)
        # /*********************************End of the code for the fast_track applicants.********************************************************************/

        self.quote_button = self.driver.find_element_by_xpath("//a[@href= '#quotes' ] ")
        self.quote_button.click()
        time.sleep(self.seconds)

        if len(self.accomodation_provider) == 0:
            self.accomodation_provider.append(self.driver.find_element_by_xpath
                                              ('//app-inline-edit[@label = "Accomodation Provider"]').text)
        else:
            self.accomodation_provider.append(self.driver.find_element_by_xpath
                                              ('//app-inline-edit[@label = "Accomodation Provider"]').text)

        if len(self.start_date) == 0:
            self.start_date.append(self.driver.find_element_by_xpath
                                   ('//app-inline-edit[@label = "Tenancy Start Date"]').text)
        else:
            self.start_date.append(self.driver.find_element_by_xpath
                                   ('//app-inline-edit[@label = "Tenancy Start Date"]').text)

        self.driver.back()
        return self.nationality, self.accomodation_provider, self.start_date, self.status, self.sales_source, self.affiliates, self.fast_track


# Using Pandas in order to edit and reshape the excel inputted and exporting it as a new .xlsx file
def excel_edit(xlsx_file):
    df = pd.read_excel(xlsx_file, engine='openpyxl', skiprows=5)

    df.columns = ["Date", "Source", "Type", "Reference", "Currency", "Debit_(Source)", "Credit_(Source)", "Debit_(GBP)",
                  "Credit_(GBP)", "Running_Balance"]

    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    df = df.loc[~df["Source"].isna()]

    df["Type"] = np.where(df["Type"].str.contains("Student", na=False), "Student", "Working Professional")
    df['Date'] = pd.to_datetime(df["Date"], format="%Y-%m-%d").dt.strftime('%d-%m-%Y')

    new_df = df.groupby(["Type", "Reference", "Date"]).agg(
        {
            "Credit_(GBP)": "sum",
            "Reference": "count"
        }
    )
    return new_df


# Converting the data from the .xlsx file exported above to a list in order to add the data extracted from the robot
def df_to_list(data):
    reference_list = []
    dataframe_to_list = data.values.tolist()
    index_list = data.index.values
    ref = []

    for i in range(0, len(index_list)):
        ref.append(list(index_list[i]))

    for i in range(0, len(ref)):
        reference_list.append(ref[i])

    for i in range(0, len(dataframe_to_list)):
        for j in range(0, len(dataframe_to_list[i])):
            reference_list[i].append(dataframe_to_list[i][j])
            # print(reference_list)
    return reference_list


# Merging the data from the .xlsx file and the robot to a list
def merge_lists(reference_list, nationality, accommodation_provider, start_date, status, sales_source, affiliates,
                fast_track):
    final_list = reference_list
    x = 0
    try:
        print(len(reference_list))
        for i in range(0, len(reference_list)):
            final_list[i].append(nationality[i])
            final_list[i].append(accommodation_provider[i])
            final_list[i].append(start_date[i])
            final_list[i].append(status[i])
            final_list[i].append(sales_source[i])
            if sales_source[i] == "Affiliates":
                final_list[i].append(affiliates[x])
                x += 1
            else:
                final_list[i].append("")
            final_list[i].append(fast_track[i])
        print(f"{i}, ", len(final_list[i]))
    except Exception as e:
        print(e)
        print(f"{i}, {final_list[i]}")
    return final_list


# Using openpyxl I'm converting the merged list to the final .xlsx and also adding some graph and also styling it.
def list_to_excel(final_list,
                  duplicates):
    book = Workbook()
    sheet = book.active
    sheet.title = "You're welcome"
    no_full = 0
    no_plan = 0
    first_time = 0
    renewed = 0
    next_cell = 0
    no_student = 1
    no_working = 1
    row_cell = 12
    col_cell = 15
    rows = 1
    dup_cell = 0
    normal = 0
    fast = 0
    months = {'January': 0, 'February': 0, 'March': 0, 'April': 0, 'May': 0, 'June': 0, 'July': 0, 'August': 0,
              'September': 0, 'October': 0, 'November': 0, 'December': 0}
    months_list = []
    header = [
        ["Type of Applicant", "Reference Number", "Payment Day", "Amount Paid", "Type of Payment", "Country of Origin",
         "Accommodation Provider", "Start Date", "Status", "Sales Source", "Affiliates", "Fast Track"]]
    font = Font(bold=True)
    border = Border(left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                    )

    # Adding a Header
    for row in header:
        sheet.append(row)
    for cell in sheet["1:1"]:
        cell.font = font
        cell.border = border
        # Inserting the merged lists row by row
    for i in range(0, len(final_list)):
        rows = i
        for j in range(0, len(final_list[i])):
            sheet.cell(row=i + 2, column=j + 1).value = final_list[i][j]

    print("\n\n", sheet.max_row, "\n\n")

    sheet.cell(row=2, column=col_cell - 1, value="Types of Applicants")

    # Calculating the amount of students and also the total amount they paid
    for i in range(2, sheet.max_row):
        if sheet[f"A{i}"].value == sheet[f"A{i + 1}"].value:
            no_student += 1
        else:
            sheet.cell(row=3, column=col_cell - 1, value=sheet[f"A{i}"].value)
            sheet.cell(row=3, column=col_cell, value=f"=SUM(D1:D{i})")
            break

    # Calculating the amount of working professionals and the total amount they paid
    for i in range(2,
                   sheet.max_row + 1):
        if sheet[f"A{i}"].value == "Working Professional":
            if sheet[f"A{i + 1}"].value is None:
                sheet.cell(row=4, column=col_cell - 1, value=sheet[f"A{i}"].value)
                sheet.cell(row=4, column=col_cell, value=f"=SUM(D{i})")
            else:
                for j in range(i, sheet.max_row + 1):
                    if sheet[f"A{j}"].value == sheet[f"A{j + 1}"].value:
                        no_working += 1
                    else:
                        sheet.cell(row=4, column=col_cell - 1, value=sheet[f"A{j}"].value)
                        sheet.cell(row=4, column=col_cell, value=f"=SUM(D{i}:D{j})")
                        break
                break

    # Creating a sum for the sums above (Students + Working Professionals)
    sheet.cell(row=5, column=col_cell - 1,
               value="Total")
    sheet.cell(row=5, column=col_cell, value=f"=SUM(O3:O4)")

    # Calculating the amount of people that paid in Full and the amount of people that paid in a Plan
    for i in range(2,
                   rows + 3):
        if sheet[f"E{i}"].value == 1:
            sheet.cell(row=i, column=5, value="Full")
            no_full += 1
            full_cell = i
        elif sheet[f"E{i}"].value > 1:
            sheet.cell(row=i, column=5, value="Payment Plan")
            no_plan += 1
            plan_cell = i

    sheet.cell(row=7, column=col_cell - 1, value="Type of Payment")
    sheet.cell(row=8, column=col_cell - 1, value=sheet[f"E{full_cell}"].value)
    sheet.cell(row=8, column=col_cell, value=no_full)
    sheet.cell(row=9, column=col_cell - 1, value=sheet[f"E{plan_cell}"].value)
    sheet.cell(row=9, column=col_cell, value=no_plan)

    sheet.cell(row=row_cell - 1, column=col_cell - 1, value="Months")

    # Converting the date to just the months
    for i in range(2, len(final_list) + 1):
        print(i)
        try:
            sheet[f"H{i}"].value = datetime.datetime.strptime(sheet[f"H{i}"].value, "%d %b, %Y").strftime("%B")
        except Exception as e:
            print(e)
            pass

    # Creating the data chart of the months
    for i in range(1, sheet.max_row + 1):
        if sheet[f"H{i}"].value in months:
            months[sheet[f"H{i}"].value] += 1

    for i in months:
        if months[i] > 0:
            months_list.append(i)
    short_dict = {key: value for key, value in months.items() if key in months_list}

    for key, value in short_dict.items():
        month_row_cell = row_cell + next_cell
        sheet.cell(row=month_row_cell, column=col_cell - 1, value=key)
        sheet.cell(row=month_row_cell, column=col_cell, value=value)
        next_cell += 1

    dup_row_cell = month_row_cell + next_cell + 1

    sheet.cell(row=dup_row_cell, column=col_cell - 1,
               value="Duplicate Ref. Numbers")  # Reporting the problematic reference numbers
    if len(duplicates) > 0:
        for i in range(len(duplicates)):
            dup_cell = i
            sheet.cell(row=dup_row_cell, column=col_cell, value=duplicates[i])
    else:
        sheet.cell(row=dup_row_cell, column=col_cell, value="None")

    # Calculating the amount of people that are First time Customer and Renewed Customers
    for i in range(2, rows + 3):
        if sheet[f"I{i}"].value == "First time customer":
            first_time += 1
            first_time_cell = i
        elif sheet[f"I{i}"].value == "Renewed":
            renewed += 1
            renewed_cell = i

    status_row_cell = dup_row_cell + dup_cell + 2

    sheet.cell(row=status_row_cell, column=col_cell - 1, value="Status")
    if first_time > 0:
        sheet.cell(row=status_row_cell + 1, column=col_cell - 1, value=sheet[f"I{first_time_cell}"].value)
        sheet.cell(row=status_row_cell + 1, column=col_cell, value=first_time)
    elif first_time == 0:
        sheet.cell(row=status_row_cell + 1, column=col_cell - 1, value="First time customer")
        sheet.cell(row=status_row_cell + 1, column=col_cell, value=first_time)
    if renewed > 0:
        sheet.cell(row=status_row_cell + 2, column=col_cell - 1, value=sheet[f"I{renewed_cell}"].value)
        sheet.cell(row=status_row_cell + 2, column=col_cell, value=renewed)
    elif renewed == 0:
        sheet.cell(row=status_row_cell + 2, column=col_cell - 1, value="Renewed")
        sheet.cell(row=status_row_cell + 2, column=col_cell, value=renewed)

    for i in range(2, sheet.max_row):
        if sheet[f"L{i}"].value == "Normal":
            normal += 1
            normal_cell = i
        elif sheet[f"L{i}"].value == "Fast Track":
            fast += 1
            fast_cell = i

    row_cell = status_row_cell + 4

    sheet.cell(row=row_cell, column=col_cell - 1, value="Fast Track")
    if normal > 0:
        sheet.cell(row=row_cell + 1, column=col_cell - 1, value=sheet[f"I{normal_cell}"].value)
        sheet.cell(row=row_cell + 1, column=col_cell, value=normal)
    elif normal == 0:
        sheet.cell(row=row_cell + 1, column=col_cell - 1, value="Normal")
        sheet.cell(row=row_cell + 1, column=col_cell, value=normal)
    if fast > 0:
        sheet.cell(row=row_cell + 2, column=col_cell - 1, value=sheet[f"I{fast_cell}"].value)
        sheet.cell(row=row_cell + 2, column=col_cell, value=fast)
    elif fast == 0:
        sheet.cell(row=row_cell + 2, column=col_cell - 1, value="Fast")
        sheet.cell(row=row_cell + 2, column=col_cell, value=fast)

    for row in range(1, sheet.max_row + 1):
        for column in "NO":
            cell_name = "{}{}".format(column, row)
            if sheet[cell_name].value is not None:
                sheet[cell_name].border = border

    # Creating a Pie chart for the Types of applicants
    type_pie = PieChart()
    labels = Reference(sheet, min_col=col_cell - 1, min_row=3, max_row=4)
    data = Reference(sheet, min_col=col_cell, min_row=2, max_row=4)
    type_pie.add_data(data, titles_from_data=True)
    type_pie.set_categories(labels)
    type_pie.title = "Students: %s, Working Professionals: %s" % (no_student, no_working)
    slice = DataPoint(idx=0, explosion=40)
    type_pie.series[0].data_points = [slice]
    sheet.add_chart(type_pie, "R2")

    payment_pie = PieChart()
    labels = Reference(sheet, min_col=col_cell - 1, min_row=8, max_row=9)
    data = Reference(sheet, min_col=col_cell, min_row=7, max_row=9)
    payment_pie.add_data(data, titles_from_data=True)
    payment_pie.set_categories(labels)
    payment_pie.title = "Full: %s, Payment Plan: %s" % (no_full, no_plan)
    slice = DataPoint(idx=0, explosion=15)
    payment_pie.series[0].data_points = [slice]
    sheet.add_chart(payment_pie, "R20")

    # Creating a chart for the Months based on the Start Date
    months_pie = PieChart()
    labels = Reference(sheet, min_col=col_cell - 1, min_row=month_row_cell - 1, max_row=month_row_cell)
    data = Reference(sheet, min_col=col_cell, min_row=month_row_cell - 2, max_row=month_row_cell)
    months_pie.add_data(data, titles_from_data=True)
    months_pie.set_categories(labels)
    months_pie.title = "Number of Applicants by Month"
    slice = DataPoint(idx=0, explosion=15)
    months_pie.series[0].data_points = [slice]
    sheet.add_chart(months_pie, "AA20")

    # Creating a chart for the Months based on the Status
    status_pie = PieChart()
    labels = Reference(sheet, min_col=col_cell - 1, min_row=status_row_cell + 1, max_row=status_row_cell + 2)
    data = Reference(sheet, min_col=col_cell, min_row=status_row_cell, max_row=status_row_cell + 2)
    status_pie.add_data(data, titles_from_data=True)
    status_pie.set_categories(labels)
    status_pie.title = "Number of Applicants by Status"
    slice = DataPoint(idx=0, explosion=15)
    status_pie.series[0].data_points = [slice]
    sheet.add_chart(status_pie, "AA2")

    # Checking the System the program is running and exporting to the Desktop accordingly
    os_base = platform.system()
    if os_base != "Darwin":
        desktop_dir = os.environ["USERPROFILE"] + "\\Desktop\\Customer Report\\"
        today = date.today().strftime("%d_%m_%y")
        try:
            book.save(filename=desktop_dir + f'output{today}.xlsx')
        except OSError as e:
            os.makedir(desktop_dir)
            book.save(filename=desktop_dir + f'output{today}.xlsx')
            
    else:
        desktop_dir = os.environ["HOME"] + "/Desktop/Customer Report/"
        today = date.today().strftime("_%d_%m_%y")
        try:
            book.save(filename=desktop_dir + f'output{today}.xlsx')
        except OSError as e:
            os.makedir(desktop_dir)
            book.save(filename=desktop_dir + f'output{today}.xlsx')
            
    return book


# Executing the code
def main():
    # Start Time.
    start = time.time()
    # Start gui for file selection. /* I need to create a credentials function  */
    excel_file = Gui()
    # Start Pandas for excel processing.
    df = excel_edit(str(excel_file.output))
    # Output the Pivot Table of the Excel file.
    df.to_excel("../data/data_table.xlsx")
    # Combine the Pivot Table to a Python List.
    ref_list = df_to_list(df)
    with open("../Credentials/credentials.ko", 'r') as cred_json:
        data = json.load(cred_json)
        for cred in data["Credentials"]:
            email = cred['Email']
            passw = cred['Password']
    # Start the Chrome and Login to Panacea.
    applicants = Applicants(email, passw)
    # Get the data for each applicant.
    applicants.get_data(ref_list)
    # Combine all the data of each applicant to a new list.
    merged_list = merge_lists(ref_list, applicants.nationality, applicants.accomodation_provider, applicants.start_date,
                              applicants.status, applicants.sales_source, applicants.affiliates,
                              applicants.fast_track)
    # Create a new excel and export is the Desktop.
    list_to_excel(merged_list, applicants.duplicates)
    # Output the time taken and then exit the program.
    print("The program took %.2f minutes" % ((time.time() - start) / 60))


if __name__ == "__main__":
    main()
